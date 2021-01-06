import pandas as pd
import openpyxl.styles
from openpyxl.styles import Side
import openpyxl
import openpyxl.styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Border
import requests
import json
import xlwt
from lxml import etree
import win32api
import win32gui
import time
import win32con
import win32clipboard as w
from win32com.client import DispatchEx
from PIL import ImageGrab, Image, ImageDraw, ImageFont
import uuid
from io import BytesIO
import os
import datetime



session = requests.Session()
qi_url = 'https://tianqi.moji.com/weather/china/henan/chuanhui-district'
aqi_url = 'https://tianqi.moji.com/aqi/china/henan/chuanhui-district'
first_url = 'http://1.192.88.18:18111/release/getMapData'
headers = {
    'User-Agent':'Dalvik/2.1.0 (Linux; U; Android 7.1.2; LIO-AN00 Build/N2G48H)'
}

def FindWindow(chatroom):
    win = win32gui.FindWindow(None, chatroom)
    print("找到群聊窗口：%x" % win)
    if win != 0:
        win32gui.ShowWindow(win, win32con.SW_SHOWMINIMIZED)
        win32gui.ShowWindow(win, win32con.SW_SHOWNORMAL)
        win32gui.ShowWindow(win, win32con.SW_SHOW)
        win32gui.SetWindowPos(win, win32con.HWND_TOPMOST, 100, 100, 300, 300, win32con.SWP_SHOWWINDOW)
        win32gui.SetForegroundWindow(win)  # 获取控制
        time.sleep(1)
    else:
        print('请注意：找不到【%s】这个人（或群），请激活窗口！' % chatroom)
        exit()

def CloseWindow(chatroom):
    win = win32gui.FindWindow(None, chatroom)
    #print("找到关闭窗口：%x" % win)
    time.sleep(3)
    win32gui.ShowWindow(win, win32con.SW_SHOWMINIMIZED)

def ctrlV():
    win32api.keybd_event(17, 0, 0, 0)  # ctrl键位码是17
    win32api.keybd_event(86, 0, 0, 0)  # v键位码是86
    win32api.keybd_event(86, 0, win32con.KEYEVENTF_KEYUP, 0)  # 释放按键
    win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)

def altS():
    win32api.keybd_event(18, 0, 0, 0)  # Alt键位码
    win32api.keybd_event(83, 0, 0, 0)  # s键位码
    win32api.keybd_event(18, 0, win32con.KEYEVENTF_KEYUP, 0)  # 释放按键
    win32api.keybd_event(83, 0, win32con.KEYEVENTF_KEYUP, 0)

def send():
    ctrlV()
    altS()

def qi():
    url = first_url
    data={
        'cityCode': '411600',
        'cityCodeArrays': 'hnAll',
    }
    response = session.post(url=url, data=data,headers=headers,timeout=10).text
    return response

def save_excel():
    book = xlwt.Workbook()
    sheet = book.add_sheet("Sheet1")
    n = 1
    sheet.write(0, 0, '排名1')
    sheet.write(0,1,'站点1')
    sheet.write(0,2,'PM10(μg/m3)')
    sheet.write(0, 3, '排名2')
    sheet.write(0, 4, '站点2')
    sheet.write(0, 5, 'PM2.5(μg/m3)')
    sheet.write(0, 6, '时间')
    sheet.write(0,7,'首要污染物')
    time.sleep(5)
    l = qi()
    data = json.loads(l)
    for k in data:
        if k['stationname'] in ['市运管处','市环境监测站','周口师范','川汇区环保局']:
            sheet.write(n, 1, k['stationname'])
            if k['pm10']==0 or k['pm10']=='-' or k['pm10']=='0':
                sheet.write(n, 2, None)
            else:
                sheet.write(n, 2, k['pm10'])
            sheet.write(n, 4, k['stationname'])
            if k['pm25']==0 or k['pm25']=='-' or k['pm25']=='0':
                sheet.write(n, 5, None)
            else:
                sheet.write(n, 5, k['pm25'])
            timeArray = time.localtime(int(k["datatime"])/1000)
            otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
            sheet.write(n, 6, otherStyleTime)
            sheet.write(n, 7, k['primarypol'])
            n += 1
    book.save('1.xls')

def excel_rank():
    df = pd.read_excel('1.xls')
    df['排名1'] = df['PM10(μg/m3)'].rank(method='min',ascending=True)
    df['排名2'] = df['PM2.5(μg/m3)'].rank(method='min',ascending=True)
    df = df.sort_values(by='PM10(μg/m3)',na_position='last')
    df.reset_index(drop=True, inplace=True)
    df11 = df['排名1']
    df12 = df['站点1']
    de13 = df['PM10(μg/m3)']
    df = df.sort_values(by='PM2.5(μg/m3)',na_position='last')
    df.reset_index(drop=True, inplace=True)
    df21 = df['排名2']
    df22 = df['站点2']
    de23 = df['PM2.5(μg/m3)']
    df24 = df['时间']
    de25 = df['首要污染物']
    dff = pd.DataFrame({'排名1':df11,'站点1':df12,'PM10(μg/m3)':de13,'排名2':df21,'站点2':df22,'PM2.5(μg/m3)':de23,'时间':df24,'首要污染物':de25})
    dff.to_excel('1.xlsx',index=False)

def set_from_center(excel_rank_insert):
    fileName = excel_rank_insert
    data = openpyxl.load_workbook(fileName)
    table = data.get_sheet_by_name("Sheet1")
    nrows = table.max_row
    ncols = table.max_column
    for i in range(nrows):
        for j in range(ncols):
            table.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
    data.save(fileName)

def table_border(excel_rank_insert):
    fileName = excel_rank_insert
    wb = openpyxl.load_workbook(fileName)
    ws = wb.get_sheet_by_name("Sheet1")
    nrows = ws.max_row
    ncols = ws.max_column
    for i in range(nrows):
        for j in range(ncols):
            border = Border(left=Side(border_style='thin', color ='000000'),
            right = Side(border_style='thin', color ='000000'),
            top = Side(border_style='thin', color ='000000'),
            bottom = Side(border_style='thin', color ='000000'))
            ws.cell(i+1, j+1).border = border
    wb.save(fileName)

def table_font(ls):
    xl = qixiang()
    wb = openpyxl.load_workbook(ls)
    sheet = wb["Sheet1"]
    nrows = sheet.max_row  # 获得行数
    ncols = sheet.max_column
    for i in range(nrows):
        for j in range(ncols):
            if sheet.cell(row=i + 1, column=j + 1).value:
                pass
            else:
                sheet.cell(row=i + 1, column=j + 1).value = '-'
    sheet.insert_rows(1)
    sheet.insert_rows(1)
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    for i in range(1,9):
        sheet.row_dimensions[i].height = 20
    sheet.cell(1, 1).value = sheet.cell(4, 7).value[5:7]+'月'+sheet.cell(4, 7).value[8:10]+'日'+sheet.cell(4, 7).value[11:13]+'时国控站点小时数据'
    sheet.cell(2, 1).value = '温度：'+xl[2]+'℃'+'      '+'湿度：'+xl[0]+'      '+'风向风级：'+xl[1]
    sheet.cell(3, 1).value = '排名'
    sheet.cell(3, 2).value = '站点'
    sheet.cell(3, 4).value = '排名'
    sheet.cell(3, 5).value = '站点'
    for i in range(4,8):
        if sheet.cell(i, 5).value == '周口师范':
            shouyao = sheet.cell(i, 8).value
            if shouyao == '-':
                shouyao = '无'
    sheet.cell(8, 1).value = '首要污染物：'+shouyao
    sheet["A1"].font = Font(size = 12,bold = True,color = "FF0000")
    sheet["A2"].font = Font(size=10, bold=True, color="000000")
    sheet["A8"].font = Font(size=10, bold=True, color="000000")
    sheet.merge_cells('A1:F1')
    sheet.merge_cells('A2:F2')
    sheet.merge_cells('A8:F8')
    wb.save(ls)
    set_from_center(ls)
    table_border(ls)

def qixiang():
    response = session.get(url=qi_url, headers=headers)
    response.encoding = "utf-8"
    res = response.text
    html = etree.HTML(res)
    shidu = html.xpath('//div[@class="wea_about clearfix"]/span/text()')[0][3:]
    fengji = html.xpath('//div[@class="wea_about clearfix"]/em/text()')[0]
    wendu = html.xpath('//div[@class="wea_weather clearfix"]/em/text()')[0]
    return shidu,fengji,wendu

def excel_catch_screen(excel_rank_insert,img_name=False):
    excel = DispatchEx("Excel.Application")
    excel.Visible = True
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(excel_rank_insert)
    ws = wb.Sheets("Sheet1")
    ws.Range("A1:F8").CopyPicture()
    ws.Paste()
    name = str(uuid.uuid4())
    new_shape_name = name[:6]
    excel.Selection.ShapeRange.Name = new_shape_name
    ws.Shapes(new_shape_name).Copy()  # 选择图片
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    if not img_name:
        img_name = r'pic\name' + ".png"
    img.save(img_name)  # 保存图片
    wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
    excel.Quit()  # 退出excel

# 获取发送的文件
def get_file(image):
    file_name = os.listdir(image)
    return file_name

# 图片转化成BPM格式图片
def paste_img(file_img):
    # 把图片写入image变量中
    # 用open函数处理后，图像对象的模式都是 RGB
    image = Image.open(file_img)

    # 声明output字节对象
    output = BytesIO()

    # 用BMP (Bitmap) 格式存储
    # 这里是位图，然后用output字节对象来存储
    image.save(output, 'BMP')

    # BMP图片有14字节的header，需要额外去除
    data = output.getvalue()[14:]

    # 关闭
    output.close()
    # DIB: 设备无关位图(device-independent bitmap)，名如其意
    # BMP的图片有时也会以.DIB和.RLE作扩展名
    # 设置好剪贴板的数据格式，再传入对应格式的数据，才能正确向剪贴板写入数据
    setImage(data)

# 把图片放到剪切板
def setImage(data):  # 写入剪切板

    w.OpenClipboard()
    try:
        # Unicode tests
        w.EmptyClipboard()
        w.SetClipboardData(win32con.CF_DIB, data)
    except:
        traceback.print_exc()
    finally:
        w.CloseClipboard()


def blend_two_images():
    img1 = Image.open(r"pic\name.png")
    img1 = img1.resize((640, 930))
    img1 = img1.convert('RGBA')

    img2 = Image.open(r"pic\ls.png")
    img2 = img2.resize((640, 930))
    img2 = img2.convert('RGBA')

    img = Image.blend(img1, img2, 0.6)
    img.show()
    img.save(r"pic\blend.png")

# 复制文本到剪切板
def setText(aString):
    w.OpenClipboard()
    w.EmptyClipboard()
    w.SetClipboardData(win32con.CF_UNICODETEXT, aString)
    w.CloseClipboard()


def send_text(name,l):
    FindWindow(name)
    setText(l)
    send()
    CloseWindow(name)

def send(name):
    save_excel()
    excel_rank()
    ls = '1.xlsx'
    xl = r'D:\Program Files\pycharm\微信自动发送数据\demo\1.xlsx'
    table_font(ls)
    excel_catch_screen(xl)
    FindWindow(name)
    ctrlV()
    altS()
    CloseWindow(name)

if __name__ == '__main__':
    f = open('time_file/datafile.txt')
    print('时间文件获取成功')
    time_list = f.read()
    name = input('请输入要发送的微信名称：')
    print('微信名称获取成功')
    while True:
        try:
            now_time = datetime.datetime.strftime(datetime.datetime.now(),'%H:%M')
            if now_time in time_list:
                print(now_time)
                send(name)
            time.sleep(60)
        except:
            send_text(name,'数据异常')
            time.sleep(60)

# send('王彦军')


# sessions = requests.Session()
# url = 'http://222.143.24.250:100/airgis_dp/messagebroker/amf'
#
# headerss = {
#     'Host': '222.143.24.250:100',
#     'Connection': 'keep-alive',
#     'Content-Length': '2717',
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
#     'Content-Type': 'application/x-amf',
#     'Accept': '*/*',
#     'Origin': 'http://222.143.24.250:100',
#     'X-Requested-With': 'ShockwaveFlash/32.0.0.465',
#     'Referer': 'http://222.143.24.250:100/flex1/index.swf/[[DYNAMIC]]/5',
#     'Accept-Encoding': 'gzip, deflate',
#     'Accept-Language': 'zh-CN,zh;q=0.9',
#     'Cookie': 'JSESSIONID=DDEC9E744D83E9BC99805A846A0C2D18',
# }
#
# res = sessions.post(url,headers = headerss).text
# print(res)