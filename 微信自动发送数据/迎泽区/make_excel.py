import pandas as pd
import openpyxl.styles
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill
import openpyxl
import openpyxl.styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import xlrd
import xlwt
from pandas import read_csv
import glob
from openpyxl.styles import Color
from openpyxl.formatting.rule import Rule
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule




# excel根据字段排名
def excel_rank(excel_file_dir,excel_filenew_dir,rank_name):
    df = pd.read_excel(excel_file_dir)
    df['排名'] = df[rank_name].rank(method='min', ascending=True)
    df['综指5排名'] = df['综合指数'].rank(method='min', ascending=True)
    df['PM2.5排名'] = df['PM2.5'].rank(method='min', ascending=True)
    df['PM10排名'] = df['PM10'].rank(method='min', ascending=True)
    df = df.sort_values(by=rank_name,na_position='last')
    df.reset_index(drop=True, inplace=True)
    df.to_excel(excel_filenew_dir,index=False)


# 充填行，按照AQI
def excel_c(excel_filenew_dir,name_c,excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    for i in range(1,8):
        if sheet.cell(i, 2).value == name_c:
            n = i
    if sheet.cell(n, 10).value:
        if 50 >= float(sheet.cell(n, 10).value) >= 0:
            color = '00E400'
        elif 100 >= float(sheet.cell(n, 10).value) > 50:
            color = 'FFFF00'
        elif 150 >= float(sheet.cell(n, 10).value) > 100:
            color = 'FF7E00'
        elif 200 >= float(sheet.cell(n, 10).value) > 150:
            color = 'FF0000'
        elif 300 >= float(sheet.cell(n, 10).value) > 200:
            color = '99004C'
        else:
            color = 'FFFFFF'
    else:
        color = 'FFFFFF'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,13):
        sheet.cell(n, j).fill = fille
    wb.save(excel_filerank_dir)
    zongzhi = sheet.cell(n, 3).value
    zongrank = sheet.cell(n, 14).value
    aqi = sheet.cell(n, 10).value
    aqidengji = sheet.cell(n, 12).value
    shouyao = sheet.cell(n, 11).value
    pm25nong = sheet.cell(n, 8).value
    pm10nong = sheet.cell(n, 9).value
    pm25rank = sheet.cell(n, 15).value
    pm10rank = sheet.cell(n, 16).value
    airtype = sheet.cell(n, 13).value
    wb.save(excel_filerank_dir)
    return (zongzhi,zongrank,aqi,aqidengji,shouyao, pm25nong, pm25rank, pm10nong, pm10rank,airtype)


#居中对齐，通过遍历方式实现
def set_from_center(excel_rank_insert):
    fileName = excel_rank_insert
    data = openpyxl.load_workbook(fileName)
    table = data.get_sheet_by_name("Sheet1")
    nrows = table.max_row # 获得行数
    ncols = table.max_column
    for i in range(nrows):
        for j in range(ncols):
            table.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    data.save(fileName)

def table_border(excel_rank_insert):
    wb = openpyxl.load_workbook(excel_rank_insert)
    # sheet = wb["Sheet1"]
    ws = wb.get_sheet_by_name("Sheet1")
    nrows = ws.max_row
    ncols = ws.max_column
    for i in range(nrows):
        for j in range(ncols):
            border = Border(left=Side(border_style='thin', color ='000000'),
            right = Side(border_style='thin', color ='000000'),
            top = Side(border_style='thin', color ='000000'),
            bottom = Side(border_style='thin', color ='000000'))
            ws.cell(i+1, j+1).border = border
    wb.save(excel_rank_insert)


def table_font(excel_filerank_dir,name_table,excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    nrows = sheet.max_row  # 获得行数
    ncols = sheet.max_column
    for i in range(nrows):
        for j in range(ncols):
            if sheet.cell(row=i + 1, column=j + 1).value:
                pass
            else:
                sheet.cell(row=i + 1, column=j + 1).value = '-'
    sheet.insert_rows(1)
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['K'].width = 12
    sheet.column_dimensions['C'].width = 10
    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[2].height = 15
    sheet.cell(1, 1).value = name_table
    sheet["A1"].font = Font(size = 14,bold = True,color = "FF0000")
    sheet.merge_cells('A1:L1')
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)


def table_fontjc(excel_filerank_dir,excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)

# 色阶
def color_scale(excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.conditional_formatting.add('G2:G12', ColorScaleRule(start_type='min', start_color='00FF00',end_type = 'max', end_color = 'FF0000'))
    wb.save(excel_filerank_dir)