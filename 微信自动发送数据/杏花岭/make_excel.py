import pandas as pd
import openpyxl.styles
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill
import openpyxl
import openpyxl.styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import xlrd
import xlwt
from pandas import read_csv
import glob
from openpyxl.styles import Color
from openpyxl.formatting.rule import Rule
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule




# excel根据字段排名
def excel_rank(excel_file_dir,excel_filenew_dir,rank_name):
    df = pd.read_excel(excel_file_dir)
    df['排名'] = df[rank_name].rank(method='min', ascending=True)
    df['排名1'] = df["浓度值1"].rank(method='min', ascending=True)
    df['排名2'] = df["浓度值2"].rank(method='min', ascending=True)
    df['排名3'] = df["浓度值3"].rank(method='min', ascending=True)
    df['排名4'] = df["浓度值4"].rank(method='min', ascending=True)
    df['排名5'] = df["浓度值5"].rank(method='min', ascending=True)
    df['排名6'] = df["浓度值6"].rank(method='min', ascending=True)
    df['综合指数排名6'] = df["综合指数"].rank(method='min', ascending=True)
    df = df.sort_values(by=rank_name,ascending=True,na_position='last')
    df.reset_index(drop=True, inplace=True)
    df.to_excel(excel_filenew_dir,index=False)


# 充填行，按照AQI
def excel_c(excel_filenew_dir,name_c,excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    for i in range(1,10):
        if sheet.cell(i, 2).value == name_c:
            n = i
    if sheet.cell(n, 15).value:
        if 4 >= float(sheet.cell(n, 15).value) >= 0:
            color = '00E400'
        elif 6 >= float(sheet.cell(n, 15).value) > 4:
            color = 'FFFF00'
        elif 8 >= float(sheet.cell(n, 15).value) > 6:
            color = 'FF7E00'
        elif 10 >= float(sheet.cell(n, 15).value) > 8:
            color = 'FF0000'
        elif float(sheet.cell(n, 15).value) > 10:
            color = '99004C'
        else:
            color = 'FFFFFF'
    else:
        color = 'FFFFFF'
    print(sheet.cell(n, 15).value)
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,16):
        sheet.cell(n, j).fill = fille
    zongzhi = sheet.cell(n, 15).value
    zongrank = sheet.cell(n, 16).value
    han = {1: '一', 2: '二', 3: '三', 4: '四', 5: '五', 6: '六', 7: '七', 8: '八', 9: '九', 10: '十', 11: '十一'}
    aqirankh = han[zongrank]
    x = 0
    for j in range(1, 12):
        if sheet.cell(j, 16).value == zongrank:
            x += 1
    if x >= 2:
        aqirank = '并列第' + str(aqirankh)
    else:
        aqirank = '第' + str(aqirankh)

    wb.save(excel_filerank_dir)
    return (zongzhi, zongrank,aqirank)



#居中对齐，通过遍历方式实现
def set_from_center(excel_rank_insert):
    fileName = excel_rank_insert
    data = openpyxl.load_workbook(fileName)
    table = data.get_sheet_by_name("Sheet1")
    nrows = table.max_row # 获得行数
    ncols = table.max_column
    for i in range(nrows):
        for j in range(ncols):
            table.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    data.save(fileName)

def table_border(excel_rank_insert):
    wb = openpyxl.load_workbook(excel_rank_insert)
    # sheet = wb["Sheet1"]
    ws = wb.get_sheet_by_name("Sheet1")
    nrows = ws.max_row
    ncols = ws.max_column
    for i in range(nrows):
        for j in range(ncols):
            border = Border(left=Side(border_style='thin', color ='000000'),
            right = Side(border_style='thin', color ='000000'),
            top = Side(border_style='thin', color ='000000'),
            bottom = Side(border_style='thin', color ='000000'))
            ws.cell(i+1, j+1).border = border
    wb.save(excel_rank_insert)


def table_font(excel_filerank_dir,name_table,excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    nrows = sheet.max_row  # 获得行数
    ncols = sheet.max_column
    for i in range(nrows):
        for j in range(ncols):
            if sheet.cell(row=i + 1, column=j + 1).value:
                pass
            else:
                sheet.cell(row=i + 1, column=j + 1).value = '-'
    sheet.insert_rows(1)
    sheet.insert_rows(1)
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['K'].width = 12
    sheet.column_dimensions['C'].width = 10
    sheet.row_dimensions[1].height = 20
    sheet.cell(1, 1).value = name_table
    sheet.cell(2, 1).value = "排名"
    sheet.cell(2, 2).value = "点位"
    sheet.cell(2, 15).value = "综合指数"
    sheet["A1"].font = Font(size = 14,bold = True,color = "FF0000")
    sheet.merge_cells('A1:O1')
    sheet.merge_cells('A2:A3')
    sheet.merge_cells('B2:B3')
    sheet.merge_cells('O2:O3')
    sheet.cell(2, 3).value = 'SO2'
    sheet.merge_cells('C2:D2')
    sheet.cell(2, 5).value = 'NO2'
    sheet.merge_cells('E2:F2')
    sheet.cell(2, 7).value = "CO"
    sheet.merge_cells('G2:H2')
    sheet.cell(2, 9).value = "O3_8h"
    sheet.merge_cells('I2:J2')
    sheet.cell(2, 11).value = "PM2.5"
    sheet.merge_cells('K2:L2')
    sheet.cell(2, 13).value = "PM10"
    sheet.merge_cells('M2:N2')
    for i in range(3,15,2):
        sheet.cell(3, i).value = "浓度值"
    for j in range(4,15,2):
        sheet.cell(3, j).value = "排名"
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)

def table_fontjc(excel_filerank_dir,excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)

# 色阶
def color_scale(excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.conditional_formatting.add('G2:G12', ColorScaleRule(start_type='min', start_color='00FF00',end_type = 'max', end_color = 'FF0000'))
    wb.save(excel_filerank_dir)