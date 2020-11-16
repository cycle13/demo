import pandas as pd
import openpyxl.styles
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill
import openpyxl
import openpyxl.styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font,colors
import xlrd
import xlwt
from pandas import read_csv
import glob




# excel根据字段排名
def excel_rank(excel_file_dir,excel_filenew_dir,rank_name):
    df = pd.read_excel(excel_file_dir)
    df['排名'] = df[rank_name].rank(method='min',ascending=True)
    df = df.sort_values(by=rank_name)
    df.reset_index(drop=True, inplace=True)
    df.to_excel(excel_filenew_dir,index=False)

# excel根据字段排名
def excel_rank_rb(excel_file_dir,excel_filenew_dir,rank_name):
    df = pd.read_excel(excel_file_dir)
    df = df.sort_values(by=rank_name,ascending=True)
    # df[excel_add_name] = df[rank_name].rank(method='min', ascending=False)
    df.reset_index(drop=True, inplace=True)
    df.to_excel(excel_filenew_dir,index=False)

def excel_rank_str(excel_file_dir,excel_filenew_dir,rank_name):
    df = pd.read_excel(excel_file_dir)
    df['排名'] = df[rank_name].rank(method='min', ascending=True)
    df['AQI排名'] = df['AQI'].rank(method='min', ascending=True)
    df = df.sort_values(by=rank_name)
    df.reset_index(drop=True, inplace=True)
    df.to_excel(excel_filenew_dir,index=False)

# 充填行，按照AQI
def excel_c(excel_filenew_dir,name_c,excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    for i in range(1,9):
        if sheet.cell(i, 2).value == name_c:
            n = i
    if 50 >= sheet.cell(n, 8).value >= 0:
        color = '00E400'
    elif 100 >= sheet.cell(n, 8).value > 50:
        color = 'FFFF00'
    elif 150 >= sheet.cell(n, 8).value > 100:
        color = 'FF7E00'
    elif 200 >= sheet.cell(n, 8).value > 150:
        color = 'FF0000'
    elif 300 >= sheet.cell(n, 8).value > 200:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,10):
        sheet.cell(n, j).fill = fille
    wb.save(excel_filerank_dir)


def excel_bz_six(excel_filenew_dir,name_c,excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    for i in range(1,8):
        if sheet.cell(i, 2).value == name_c:
            n = i
    if 50 >= sheet.cell(n, 11).value >= 0:
        color = '00E400'
    elif 100 >= sheet.cell(n, 11).value > 50:
        color = 'FFFF00'
    elif 150 >= sheet.cell(n, 11).value > 100:
        color = 'FF7E00'
    elif 200 >= sheet.cell(n, 11).value > 150:
        color = 'FF0000'
    elif 300 >= sheet.cell(n, 11).value > 200:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,13):
        sheet.cell(n, j).fill = fille
    wb.save(excel_filerank_dir)


def excel_bz_any(excel_filenew_dir,name_c,excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    for i in range(1,13):
        if sheet.cell(i, 2).value == name_c:
            n = i
    if 50 >= sheet.cell(n, 9).value >= 0:
        color = '00E400'
    elif 100 >= sheet.cell(n, 9).value > 50:
        color = 'FFFF00'
    elif 150 >= sheet.cell(n, 9).value > 100:
        color = 'FF7E00'
    elif 200 >= sheet.cell(n, 9).value > 150:
        color = 'FF0000'
    elif 300 >= sheet.cell(n, 9).value > 200:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,12):
        sheet.cell(n, j).fill = fille
    wb.save(excel_filerank_dir)
    aqi = sheet.cell(n, 9).value
    aqidengji = sheet.cell(n, 11).value
    shouyao = sheet.cell(n, 10).value
    aqirank = sheet.cell(n, 12).value
    return (aqi, aqidengji, shouyao, aqirank)



def excel_bz_hour(excel_filenew_dir,name_c,excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    for i in range(1,8):
        if sheet.cell(i, 2).value == name_c:
            n = i
    if 50 >= sheet.cell(n, 9).value >= 0:
        color = '00E400'
    elif 100 >= sheet.cell(n, 9).value > 50:
        color = 'FFFF00'
    elif 150 >= sheet.cell(n, 9).value > 100:
        color = 'FF7E00'
    elif 200 >= sheet.cell(n, 9).value > 150:
        color = 'FF0000'
    elif 300 >= sheet.cell(n, 9).value > 200:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,12):
        sheet.cell(n, j).fill = fille
    wb.save(excel_filerank_dir)
    aqi = sheet.cell(n, 9).value
    aqidengji = sheet.cell(n, 11).value
    shouyao = sheet.cell(n, 10).value
    aqirank = sheet.cell(n, 12).value
    return (aqi, aqidengji, shouyao, aqirank)


def excel_add(excel_filenew_dir,name_c,excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    for i in range(1,13):
        if sheet.cell(i, 2).value == name_c:
            n = i
    if 50 >= sheet.cell(n, 9).value >= 0:
        color = '00E400'
    elif 100 >= sheet.cell(n, 9).value > 50:
        color = 'FFFF00'
    elif 150 >= sheet.cell(n, 9).value > 100:
        color = 'FF7E00'
    elif 200 >= sheet.cell(n, 9).value > 150:
        color = 'FF0000'
    elif 300 >= sheet.cell(n, 9).value > 200:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,13):
        sheet.cell(n, j).fill = fille
    wb.save(excel_filerank_dir)

def col_num(sheet,name_c):
    n = 0
    for i in range(1, 13):
        if sheet.cell(i, 1).value == name_c:
            n = i
    return n


def aqi_color(color_num):
    if 50 >= color_num >= 0:
        color = '00E400'
    elif 100 >= color_num > 50:
        color = 'FFFF00'
    elif 150 >= color_num > 100:
        color = 'FF7E00'
    elif 200 >= color_num > 150:
        color = 'FF0000'
    elif 300 >= color_num > 200:
        color = '99004C'
    else:
        color = '7E0023'
    return color


def excel_c_hour(excel_filenew_dir1,name_c1,name_c2,excel_filerank_dir1):
    wb = openpyxl.load_workbook(excel_filenew_dir1)
    sheet = wb["Sheet1"]
    n = col_num(sheet,name_c1)
    for j in range(1,5):
        sheet.cell(n, j).font = Font(bold=True, color="FF0000")

    m = col_num(sheet, name_c2)
    for j in range(1, 5):
        sheet.cell(m, j).font = Font(bold=True, color="FF0000")

    for j in range(2,11):
        c = sheet.cell(j, 2).value
        fille = PatternFill("solid", fgColor=aqi_color(c))
        sheet.cell(j, 2).fill = fille
    wb.save(excel_filerank_dir1)


#居中对齐，通过遍历方式实现
def set_from_center(excel_rank_insert):
    fileName = excel_rank_insert
    data = openpyxl.load_workbook(fileName)
    table = data.get_sheet_by_name("Sheet1")
    nrows = table.max_row # 获得行数
    ncols = table.max_column
    for i in range(nrows):
        for j in range(ncols):
            table.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    data.save(fileName)

def table_border(excel_rank_insert):
    wb = openpyxl.load_workbook(excel_rank_insert)
    # sheet = wb["Sheet1"]
    ws = wb.get_sheet_by_name("Sheet1")
    nrows = ws.max_row
    ncols = ws.max_column
    for i in range(nrows):
        for j in range(ncols):
            border = Border(left=Side(border_style='thin', color ='000000'),
            right = Side(border_style='thin', color ='000000'),
            top = Side(border_style='thin', color ='000000'),
            bottom = Side(border_style='thin', color ='000000'))
            ws.cell(i+1, j+1).border = border
    wb.save(excel_rank_insert)


def table_font(excel_filerank_dir,name_table,excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['I'].width = 12
    # sheet.column_dimensions['D'].width = 15
    sheet.row_dimensions[1].height = 20
    sheet.cell(1, 1).value = name_table
    sheet["A1"].font = Font(size = 16,bold = True,color = "FF0000")
    sheet.merge_cells('A1:I1')
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)


def table_font_rb(excel_filenew_dir,name_table,excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.row_dimensions[1].height = 20
    sheet.cell(1, 1).value = name_table
    sheet["A1"].font = Font(size = 14,bold = True,color = "FF0000")
    sheet.merge_cells('A1:C1')
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)


def table_font_db(excel_filenew_dir,name_table,excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    # sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    # sheet.column_dimensions['E'].width = 20
    sheet.row_dimensions[1].height = 30
    sheet.cell(1, 1).value = name_table
    sheet["A1"].font = Font(size = 10,bold = True,color = "FF0000")
    sheet.merge_cells('A1:C1')
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)


def table_font_str(excel_filenew_dir,name_table,excel_rank_insert,excel_lable):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 18
    sheet.row_dimensions[1].height = 30
    sheet.cell(1, 1).value = name_table
    sheet["A1"].font = Font(size = 14,bold = True,color = "FF0000")
    sheet.merge_cells('A1:E1')
    sheet.cell(18, 1).value = excel_lable
    sheet["A18"].font = Font(size=10, bold=True, color="FF0000")
    sheet.merge_cells('A18:E18')
    sheet['A18'].alignment = Alignment(horizontal='left', vertical='center')
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)


def table_font_six(excel_filerank_dir1, name_table1, excel_rank_insert1):
    wb = openpyxl.load_workbook(excel_filerank_dir1)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    # sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['L'].width = 15
    sheet.row_dimensions[1].height = 30
    sheet.cell(1, 1).value = name_table1
    sheet["A1"].font = Font(size = 14,bold = True,color = "FF0000")
    sheet.merge_cells('A1:L1')
    wb.save(excel_rank_insert1)
    set_from_center(excel_rank_insert1)
    table_border(excel_rank_insert1)


def table_font_any(excel_filerank_dir2, name_table2, excel_rank_insert2):
    wb = openpyxl.load_workbook(excel_filerank_dir2)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['B'].width = 15
    # sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['J'].width = 12
    sheet.row_dimensions[1].height = 20
    sheet.cell(1, 1).value = name_table2
    sheet["A1"].font = Font(size = 14,bold = True,color = "FF0000")
    sheet.merge_cells('A1:K1')
    wb.save(excel_rank_insert2)
    set_from_center(excel_rank_insert2)
    table_border(excel_rank_insert2)


def table_font_add(excel_filerank_dir2, name_table2, excel_rank_insert2):
    wb = openpyxl.load_workbook(excel_filerank_dir2)
    sheet = wb["Sheet1"]
    # sheet.insert_rows(1)
    sheet.column_dimensions['A'].width = 13
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['J'].width = 13
    sheet.column_dimensions['L'].width = 10
    # sheet.row_dimensions[1].height = 30
    sheet.cell(13, 1).value = name_table2
    sheet["A13"].font = Font(size = 10,bold = True,color = "FF0000")
    sheet.merge_cells('A13:L13')
    wb.save(excel_rank_insert2)
    set_from_center(excel_rank_insert2)
    table_border(excel_rank_insert2)


def table_font_hour(excel_filerank_dir1, name_table1, excel_rank_insert1):
    wb = openpyxl.load_workbook(excel_filerank_dir1)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['D'].width = 15
    # sheet.column_dimensions['J'].width = 13
    # sheet.column_dimensions['L'].width = 10
    # sheet.row_dimensions[1].height = 30
    sheet.cell(1, 1).value = name_table1
    sheet["A1"].font = Font(size = 12,bold = True,color = "000000")
    sheet.merge_cells('A1:D1')
    sheet.cell(12, 1).value = '(数据来源于网络，未经审核，仅供参考)'
    sheet["A12"].font = Font(size=8, bold=True, color="000000")
    sheet.merge_cells('A12:D12')
    wb.save(excel_rank_insert1)
    set_from_center(excel_rank_insert1)
    table_border(excel_rank_insert1)

