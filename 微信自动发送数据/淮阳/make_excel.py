import pandas as pd
import openpyxl.styles
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill
import openpyxl
import openpyxl.styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import xlrd
import xlwt
from pandas import read_csv
import glob
from openpyxl.styles import Color
from openpyxl.formatting.rule import Rule
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule




# excel根据字段排名
def excel_rank(excel_file_dir, excel_filenew_dir1,excel_filenew_dir2,rank_name1,rank_name2,add_name1,add_name2):
    df = pd.read_excel(excel_file_dir)
    df[add_name1] = df[rank_name1].rank(method='min',ascending=True)
    df[add_name2] = df[rank_name2].rank(method='min',ascending=True)
    df = df.sort_values(by=add_name1)
    df.reset_index(drop=True, inplace=True)
    df1 = df.sort_values(by=add_name2)
    df1.reset_index(drop=True, inplace=True)
    df.to_excel(excel_filenew_dir1,index=False)
    df1.to_excel(excel_filenew_dir2, index=False)


def excel_yearrank(excel_file_dir,excel_file_dir1):
    df = pd.read_excel(excel_file_dir)
    df["CO排名"] = df["CO"].rank(method='min',ascending=True)
    df["O3排名"] = df["O3"].rank(method='min',ascending=True)
    df["SO2排名"] = df["SO2"].rank(method='min', ascending=True)
    df["NO2排名"] = df["NO2"].rank(method='min', ascending=True)
    df["PM2.5排名"] = df["PM2.5"].rank(method='min', ascending=True)
    df["PM10排名"] = df["PM10"].rank(method='min', ascending=True)
    df["综合指数排名"] = df["综合指数"].rank(method='min', ascending=True)
    df.reset_index(drop=True, inplace=True)
    df.to_excel(excel_file_dir1,index=False)


def excel_yearc(excel_filenew_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    m = 0
    for i in range(1,11):
        if sheet.cell(i, 1).value == '淮阳县':
            n = i
    color = 'FFC000'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,16):
        sheet.cell(n, j).fill = fille

    for i in range(1,11):
        if sheet.cell(i, 1).value == '太康县':
            m = i
    color = '00B0F0'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,16):
        sheet.cell(m, j).fill = fille
    wb.save(excel_filenew_dir)
    set_from_center(excel_filenew_dir)
    table_border(excel_filenew_dir)

# 充填行，按照AQI
def excel_c(excel_filenew_dir, name_c, excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0
    for i in range(1,11):
        if sheet.cell(i, 1).value == name_c:
            n = i
    if 50 >= sheet.cell(n, 9).value >= 0:
        color = '00E400'
    elif 100 >= sheet.cell(n, 9).value > 50:
        color = 'FFFF00'
    elif 150 >= sheet.cell(n, 9).value > 100:
        color = 'FF7E00'
    elif 200 >= sheet.cell(n, 9).value > 150:
        color = 'FF0000'
    elif 300 >= sheet.cell(n, 9).value > 200:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,13):
        sheet.cell(n, j).fill = fille
    pm25time = sheet.cell(n, 2).value
    pm25nong = sheet.cell(n, 8).value
    pm10nong = sheet.cell(n, 7).value
    pm25rank = sheet.cell(n, 11).value
    pm10rank = sheet.cell(n, 12).value
    # hbstation = sheet.cell(n, 13).value
    # hbpm25 = sheet.cell(n, 15).value
    # hbpm10 = sheet.cell(n, 14).value
    # zlstation = sheet.cell(n, 16).value
    # zlpm25 = sheet.cell(n, 18).value
    # zlpm10 = sheet.cell(n, 17).value
    # bgstation = sheet.cell(n, 19).value
    # bgpm25 = sheet.cell(n, 21).value
    # bgpm10 = sheet.cell(n, 20).value
    wb.save(excel_filerank_dir)
    # return (pm25time,pm25nong,pm25rank,pm10nong,pm10rank,hbstation,hbpm25,hbpm10,zlstation,zlpm25,zlpm10,bgstation,bgpm25,bgpm10)
    return (pm25time,pm25nong,pm25rank,pm10nong,pm10rank)


def excel_cleiji(excel_filenew_dir, name_c, excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filenew_dir)
    sheet = wb["Sheet1"]
    n = 0

    for i in range(1,11):
        if sheet.cell(i, 1).value == name_c:
            n = i
    if 35 >= sheet.cell(n, 8).value >= 0:
        color = '00E400'
    elif 75 >= sheet.cell(n, 8).value > 35:
        color = 'FFFF00'
    elif 115 >= sheet.cell(n, 8).value > 75:
        color = 'FF7E00'
    elif 150 >= sheet.cell(n, 8).value > 115:
        color = 'FF0000'
    elif 250 >= sheet.cell(n, 8).value > 150:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,13):
        sheet.cell(n, j).fill = fille
    pm25time = sheet.cell(n, 2).value
    pm25nong = sheet.cell(n, 8).value
    pm10nong = sheet.cell(n, 7).value
    pm25rank = sheet.cell(n, 9).value
    pm10rank = sheet.cell(n, 10).value
    wb.save(excel_filerank_dir)
    return (pm25time,pm25nong,pm25rank,pm10nong,pm10rank)


#居中对齐，通过遍历方式实现
def set_from_center(excel_rank_insert):
    fileName = excel_rank_insert
    data = openpyxl.load_workbook(fileName)
    table = data.get_sheet_by_name("Sheet1")
    nrows = table.max_row # 获得行数
    ncols = table.max_column
    for i in range(nrows):
        for j in range(ncols):
            table.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
    data.save(fileName)



def table_border(excel_rank_insert):
    fileName = excel_rank_insert
    wb = openpyxl.load_workbook(fileName)
    # sheet = wb["Sheet1"]
    ws = wb.get_sheet_by_name("Sheet1")
    nrows = ws.max_row
    ncols = ws.max_column
    for i in range(nrows):
        for j in range(ncols):
            border = Border(left=Side(border_style='thin', color ='000000'),
            right = Side(border_style='thin', color ='000000'),
            top = Side(border_style='thin', color ='000000'),
            bottom = Side(border_style='thin', color ='000000'))
            ws.cell(i+1, j+1).border = border
    wb.save(fileName)


def table_font(excel_filerank_dir, name_table, excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['J'].width = 12
    sheet.column_dimensions['K'].width = 12
    sheet.column_dimensions['L'].width = 12
    sheet.cell(1, 1).value = name_table
    sheet["A1"].font = Font(size = 16,bold = True,color = "000000")
    sheet.merge_cells('A1:L1')
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)


def table_font_year(excel_filerank_dir, name_table):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['M'].width = 10
    sheet.column_dimensions['N'].width = 10
    sheet.column_dimensions['O'].width = 12
    sheet.cell(1, 1).value = name_table
    sheet["A1"].font = Font(size = 16,bold = True,color = "000000")
    sheet.merge_cells('A1:O1')
    wb.save(excel_filerank_dir)
    set_from_center(excel_filerank_dir)
    table_border(excel_filerank_dir)


def table_fontleiji(excel_filerank_dir, name_table, excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['I'].width = 12
    sheet.column_dimensions['J'].width = 12
    sheet.cell(1, 1).value = name_table
    sheet["A1"].font = Font(size = 16,bold = True,color = "000000")
    sheet.merge_cells('A1:J1')
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)


def table_fontjc(excel_filerank_dir,excel_rank_insert):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    wb.save(excel_rank_insert)
    set_from_center(excel_rank_insert)
    table_border(excel_rank_insert)

# 色阶
def color_scale(excel_filerank_dir):
    wb = openpyxl.load_workbook(excel_filerank_dir)
    sheet = wb["Sheet1"]
    sheet.conditional_formatting.add('G2:G12', ColorScaleRule(start_type='min', start_color='00FF00',end_type = 'max', end_color = 'FF0000'))
    wb.save(excel_filerank_dir)