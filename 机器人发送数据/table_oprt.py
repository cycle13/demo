import openpyxl.styles
import openpyxl.styles
from openpyxl.styles import Side
import openpyxl
import openpyxl.styles
from openpyxl.styles import Alignment
from openpyxl.styles import  Font, Border


#居中对齐，通过遍历方式实现
def set_from_center():
    fileName = r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据累积排名插入.xlsx"
    data = openpyxl.load_workbook(fileName)
    table = data.get_sheet_by_name("Sheet1")
    nrows = table.max_row # 获得行数
    ncols = table.max_column
    for i in range(nrows):
        for j in range(ncols):
            table.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
    data.save(fileName)


def table_border():
    fileName = r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据累积排名插入.xlsx"
    wb = openpyxl.load_workbook(fileName)
    # sheet = wb["Sheet1"]
    ws = wb.get_sheet_by_name("Sheet1")
    nrows = ws.max_row
    ncols = ws.max_column
    for i in range(nrows):
        for j in range(ncols):
            border = Border(left=Side(border_style='thin', color ='000000'),
            right = Side(border_style='thin', color ='000000'),
            top = Side(border_style='thin', color ='000000'),
            bottom = Side(border_style='thin', color ='000000'))
            ws.cell(i+1, j+1).border = border
    wb.save(fileName)


def table_font(name):
    wb = openpyxl.load_workbook(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据累积排名充填.xlsx")
    sheet = wb["Sheet1"]
    sheet.insert_rows(1)
    sheet.column_dimensions['I'].width = 12
    sheet.column_dimensions['J'].width = 12
    sheet.cell(1, 1).value = name
    sheet["A1"].font = Font(size = 15,bold = True,color = "000000")
    sheet.merge_cells('A1:J1')
    wb.save(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据累积排名插入.xlsx")
    set_from_center()
    table_border()
