import win32api
import win32con
import win32gui
import pandas as pd
import time
import win32clipboard as w
import requests
import json
from win32com.client import Dispatch, DispatchEx
from PIL import ImageGrab, Image
import uuid
import xlwt
import openpyxl
import openpyxl.styles
from openpyxl.styles import PatternFill
from datetime import datetime
from pandas import Series,DataFrame
from numpy.random import randn
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import matplotlib
from io import BytesIO


def excel_c():
    wb = openpyxl.load_workbook(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名.xlsx")
    sheet = wb["Sheet1"]
    n = 0

    for i in range(1,11):
        if sheet.cell(i, 1).value == "淮阳县":
            n = i
    if 35 >= sheet.cell(n, 8).value >= 0:
        color = '00E400'
    elif 75 >= sheet.cell(n, 8).value > 35:
        color = 'FFFF00'
    elif 115 >= sheet.cell(n, 8).value > 75:
        color = 'FF7E00'
    elif 150 >= sheet.cell(n, 8).value > 115:
        color = 'FF0000'
    elif 250 >= sheet.cell(n, 8).value > 150:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,13):
        sheet.cell(n, j).fill = fille
    pm25time = sheet.cell(n, 2).value
    pm25nong = sheet.cell(n, 8).value
    pm10nong = sheet.cell(n, 7).value
    pm25rank = sheet.cell(n, 11).value
    pm10rank = sheet.cell(n, 12).value
    wb.save(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名充填.xlsx")
    return (pm25time,pm25nong,pm25rank,pm10nong,pm10rank)


excel_c()
