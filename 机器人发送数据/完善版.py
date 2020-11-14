import win32api
import win32con
import win32gui
import pandas as pd
import time
import win32clipboard as w
import requests
import json
from win32com.client import Dispatch, DispatchEx
from PIL import ImageGrab, Image
import uuid
import xlwt
import openpyxl
import openpyxl.styles
from openpyxl.styles import PatternFill
from datetime import datetime
from pandas import Series,DataFrame
from numpy.random import randn
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import matplotlib
from io import BytesIO
import os

import table_opr



session = requests.Session()
first_url = 'http://service.envicloud.cn:8082/v2/weatherhistory/ZGVTBZE0NDI5MDM0MJAZNDC=/'
first_url_day = 'http://1.192.88.18:8115/hnAqi/v1.0/api/air/dayAqi2018_county'
first_url_month = 'http://1.192.88.18:8115/hnAqi/v1.0/api/air/airreport2018_county'
first_url_year = 'http://1.192.88.18:8115/hnAqi/v1.0/api/air/airreport2018_county '
url_list = 'http://1.192.88.18:8115/hnAqi/v1.0/api/air/getCityStationDetail'
headers = {
    'User-Agent':'Dalvik/2.1.0 (Linux; U; Android 7.1.2; LIO-AN00 Build/N2G48H)'
}

def FindWindow(chatroom):
    win = win32gui.FindWindow(None, chatroom)
    print("找到群聊窗口：%x" % win)
    if win != 0:
        win32gui.ShowWindow(win, win32con.SW_SHOWMINIMIZED)
        win32gui.ShowWindow(win, win32con.SW_SHOWNORMAL)
        win32gui.ShowWindow(win, win32con.SW_SHOW)
        win32gui.SetWindowPos(win, win32con.HWND_TOPMOST, 100, 100, 300, 300, win32con.SWP_SHOWWINDOW)
        win32gui.SetForegroundWindow(win)  # 获取控制
        time.sleep(1)
    else:
        print('请注意：找不到【%s】这个人（或群），请激活窗口！' % chatroom)
        exit()

def CloseWindow(chatroom):
    win = win32gui.FindWindow(None, chatroom)
    #print("找到关闭窗口：%x" % win)
    time.sleep(3)
    win32gui.ShowWindow(win, win32con.SW_SHOWMINIMIZED)

def setText(aString):
    w.OpenClipboard()
    w.EmptyClipboard()
    w.SetClipboardData(win32con.CF_UNICODETEXT, aString)
    w.CloseClipboard()

def ctrlV():
    win32api.keybd_event(17, 0, 0, 0)  # ctrl键位码是17
    win32api.keybd_event(86, 0, 0, 0)  # v键位码是86
    win32api.keybd_event(86, 0, win32con.KEYEVENTF_KEYUP, 0)  # 释放按键
    win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)

def altS():
    win32api.keybd_event(18, 0, 0, 0)  # Alt键位码
    win32api.keybd_event(83, 0, 0, 0)  # s键位码
    win32api.keybd_event(18, 0, win32con.KEYEVENTF_KEYUP, 0)  # 释放按键
    win32api.keybd_event(83, 0, win32con.KEYEVENTF_KEYUP, 0)

def sendText(chatrooms,text):
    for chatroom in chatrooms:
        FindWindow(chatroom)
        #文字首行留空，防止带表情复制不完全
        setText(" "+text)
        time.sleep(1)
        ctrlV()
        time.sleep(1)
        altS()
        CloseWindow(chatroom)

def year(url):
    data = {
        'end':'2020-10-22',
        'sort':'asc',
        'start':'2020-01-01'
    }
    response = session.post(url= url,data=data,headers = headers).text
    print(response)
    return response

def month(url):
    data = {
        'sort':'asc',
        'month':'2020-07'
    }
    response = session.post(url= url,data=data,headers = headers).text
    return response

def day(url):
    data = {
        'sort':'asc',
        'time':'2020-09-27'
    }
    response = session.post(url= url,data=data,headers = headers).text
    return response

def real(url):
    response = session.get(url= url_list,headers = headers).text
    return response

def save_excel():
    book = xlwt.Workbook()
    sheet = book.add_sheet('淮阳县空气质量数据')
    n = 1
    sheet.write(0,0,'区县')
    sheet.write(0,1,'时间')
    sheet.write(0,2,'CO')
    sheet.write(0,3,'O3')
    sheet.write(0,4,'SO2')
    sheet.write(0,5,'NO2')
    sheet.write(0,6,'PM10')
    sheet.write(0,7,'PM2.5')
    sheet.write(0,8,'AQI')
    sheet.write(0,9,'首要污染物')
    time.sleep(5)
    l = real(url_list)
    data = json.loads(l)['detail']
    for k in data:
        if k['CITY'] in ['沈丘县','商水县','西华县','扶沟县','郸城县','淮阳县','太康县','鹿邑县','项城市']:
            sheet.write(n, 0, k['CITY'])
            try:
                sheet.write(n, 1, k["MONIDATE"].split(" ")[1])
            except:
                sheet.write(n, 1, "无")
            sheet.write(n, 2, k['CO'])
            sheet.write(n, 3, k['O3'])
            sheet.write(n, 4, k['SO2'])
            sheet.write(n, 5, k['NO2'])
            sheet.write(n, 6, k['PM10'])
            sheet.write(n, 7, k['PM25'])
            sheet.write(n, 8, k['AQI'])
            sheet.write(n, 9, k['PRIMARYPOLLUTANT'])
            n+=1
    book.save('周口市区县数据.xls')

def excel_catch_screen():
    """ 对excel的表格区域进行截图——用例：excel_catch_screen(r"E:\年周口市区县27日数据.xls", "淮阳县空气质量数据", "A1:J10")"""
    # pythoncom.CoInitialize()  # excel多线程相关

    excel = DispatchEx("Excel.Application")  # 启动excel
    excel.Visible = True  # 可视化
    excel.DisplayAlerts = False  # 是否显示警告
    wb = excel.Workbooks.Open(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名插入.xlsx")  # 打开excel
    # wb = excel.Workbooks.Open(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名充填.xlsx")  # 打开excel
    ws = wb.Sheets("Sheet1")  # 选择sheet
    ws.Range("A1:L11").CopyPicture()  # 复制图片区域
    ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置
    name = str(uuid.uuid4())  # 重命名唯一值
    new_shape_name = name[:6]
    excel.Selection.ShapeRange.Name = new_shape_name  # 将刚刚选择的Shape重命名，避免与已有图片混淆

    ws.Shapes(new_shape_name).Copy()  # 选择图片
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    # if not img_name:
    #     img_name = name + ".PNG"
    # img.save(img_name)  # 保存图片
    wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
    excel.Quit()  # 退出excel
    # pythoncom.CoUninitialize()

def excel_rank():
    df = pd.read_excel(r'D:\Program Files\pycharm\机器人发送数据\周口市区县数据.xls')
    df['PM2.5排名'] = df['PM2.5'].rank(method='min',ascending=True)
    df['PM10排名'] = df['PM10'].rank(method='min',ascending=True)
    df = df.sort_values(by='PM2.5排名')
    df.reset_index(drop=True, inplace=True)
    df1 = df.sort_values(by='PM10排名')
    df1.reset_index(drop=True, inplace=True)
    df.to_excel(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名.xlsx",index=False)
    df1.to_excel(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名PM10.xlsx", index=False)

def excel_c():
    wb = openpyxl.load_workbook(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名.xlsx")
    sheet = wb["Sheet1"]
    n = 0

    for i in range(1,11):
        if sheet.cell(i, 1).value == "淮阳县":
            n = i
    if 50 >= sheet.cell(n, 9).value >= 0:
        color = '00E400'
    elif 100 >= sheet.cell(n, 9).value > 50:
        color = 'FFFF00'
    elif 150 >= sheet.cell(n, 9).value > 100:
        color = 'FF7E00'
    elif 200 >= sheet.cell(n, 9).value > 150:
        color = 'FF0000'
    elif 300 >= sheet.cell(n, 9).value > 200:
        color = '99004C'
    else:
        color = '7E0023'
    fille=PatternFill("solid",fgColor=color)
    for j in range(1,13):
        sheet.cell(n, j).fill = fille
    pm25time = sheet.cell(n, 2).value
    pm25nong = sheet.cell(n, 8).value
    pm10nong = sheet.cell(n, 7).value
    pm25rank = sheet.cell(n, 11).value
    pm10rank = sheet.cell(n, 12).value
    wb.save(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名充填.xlsx")
    return (pm25time,pm25nong,pm25rank,pm10nong,pm10rank)

def qi(station,datatime_n,url):
    url = url + station + '/' + datatime_n
    response = session.get(url=url, headers=headers).text
    return response

def line_bar(hour_d):
    df = pd.read_excel(r'D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名充填.xlsx')
    plt.figure()
    for i in range(len(df["区县"])):
        if 35 >= df["PM2.5"][i] >= 0:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#00E400',edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#00E400')
        elif 75 >= df["PM2.5"][i] > 35:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#FFFF00', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#FFFF00')
        elif 115 >= df["PM2.5"][i] > 75:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#FF7E00', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#FF7E00')
        elif 150 >= df["PM2.5"][i] > 115:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#FF0000', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#FF0000')
        elif 250 >= df["PM2.5"][i] > 150:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#99004C', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#99004C')
        else:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#7E0023', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM2.5"][i], fc='#7E0023')
    for i in range(len(df["PM2.5"])):
        plt.text(df["区县"][i], df["PM2.5"][i] + 0.5, '%s' % round(df["PM2.5"][i], 3), ha='center', fontsize=10)
    # plt.legend()
    font1 = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=20)
    plt.title(u"周口市九区县{}时PM2.5浓度柱状图".format(hour_d), fontproperties=font1)
    # 横坐标名称
    plt.xlabel("周口市九区县")
    plt.rcParams['font.sans-serif'] = ['SimHei']
    # 纵坐标名称
    plt.ylabel("浓度μg/m3")

    # 保存图片到本地
    plt.savefig('image/pci.png')
    matplotlib.use('Agg')
    # 显示图片
    # plt.show()

def line_barpm10(hour_d):
    df = pd.read_excel(r'D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名PM10.xlsx')
    plt.figure()
    for i in range(len(df["区县"])):
        if 50 >= df["PM10"][i] >= 0:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM10"][i], fc='#00E400',edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM10"][i], fc='#00E400')
        elif 150 >= df["PM10"][i] > 50:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM10"][i], fc='#FFFF00', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM10"][i], fc='#FFFF00')
        elif 250 >= df["PM10"][i] > 150:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM10"][i], fc='#FF7E00', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM10"][i], fc='#FF7E00')
        elif 350 >= df["PM10"][i] > 250:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM10"][i], fc='#FF0000', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM10"][i], fc='#FF0000')
        elif 420 >= df["PM10"][i] > 350:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM10"][i], fc='#99004C', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM10"][i], fc='#99004C')
        else:
            if df["区县"][i] == '淮阳县':
                plt.bar(df["区县"][i], df["PM10"][i], fc='#7E0023', edgecolor='black')
            else:
                plt.bar(df["区县"][i], df["PM10"][i], fc='#7E0023')
    for i in range(len(df["PM10"])):
        plt.text(df["区县"][i], df["PM10"][i] + 0.5, '%s' % round(df["PM10"][i], 3), ha='center', fontsize=10)
    # plt.legend()
    font1 = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=20)
    plt.title(u"周口市九区县{}时PM10浓度柱状图".format(hour_d), fontproperties=font1)
    # 横坐标名称
    plt.xlabel("周口市九区县")
    plt.rcParams['font.sans-serif'] = ['SimHei']
    # 纵坐标名称
    plt.ylabel("浓度μg/m3")

    # 保存图片到本地
    plt.savefig('image/pci1.png')
    matplotlib.use('Agg')
    # 显示图片
    # plt.show()

def get_file(image):
    file_name = os.listdir(image)
    return file_name

# 图片转化成BPM格式图片
def paste_img(file_img):
    # 把图片写入image变量中
    # 用open函数处理后，图像对象的模式都是 RGB
    image = Image.open(file_img)

    # 声明output字节对象
    output = BytesIO()

    # 用BMP (Bitmap) 格式存储
    # 这里是位图，然后用output字节对象来存储
    image.save(output, 'BMP')

    # BMP图片有14字节的header，需要额外去除
    data = output.getvalue()[14:]

    # 关闭
    output.close()
    # DIB: 设备无关位图(device-independent bitmap)，名如其意
    # BMP的图片有时也会以.DIB和.RLE作扩展名
    # 设置好剪贴板的数据格式，再传入对应格式的数据，才能正确向剪贴板写入数据
    setImage(data)

# 把图片放到剪切板
def setImage(data):  # 写入剪切板

    w.OpenClipboard()
    try:
        # Unicode tests
        w.EmptyClipboard()
        w.SetClipboardData(win32con.CF_DIB, data)
    except:
        traceback.print_exc()
    finally:
        w.CloseClipboard()

def fasong():
    ctrlV()
    altS()

# 获取文件夹下的所有图片
def get_file(image):
    file_name = os.listdir(image)
    return file_name

# 每发完一次删除文件夹下的图片
def del_files(path_file):
    ls = os.listdir(path_file)
    for i in ls:
        f_path = os.path.join(path_file, i)
        # 判断是否是一个目录,若是,则递归删除
        if os.path.isdir(f_path):
            del_files(f_path)
        else:
            os.remove(f_path)


if __name__ == '__main__':
    name = "淮阳区环境攻坚群"
    while True:
        try:
            FindWindow(name)
            save_excel()
            print("已获取完数据")
            excel_rank()
            print("已对数据排名")
            l = excel_c()
            print(l)
            print("已对数据充填")
            datatime_n = datetime.strftime(pd.datetime.now(),'%Y%m%d/%H')
            m = qi("101181404", datatime_n, first_url)
            print("已获取气象数据")
            data = json.loads(m)
            try:
                if l[0][0:2] != "无":
                    if data['rcode'] == 200:
                        setText("【实时播报】：\n淮阳区{}时，PM2.5浓度为{}μg/m3，在全市9个区县中排名第{}；PM10浓度为{}μg/m3，在全市9个区县中排名第{}。当前湿度{}%，温度为{}℃，风力为{}，风向为{}。".format(l[0][0:2],l[1],l[2],l[3],l[4],data['humidity'],data['temperature'],data['windpower'],data['winddirect']))
                    else:
                        setText(
                            "【实时播报】：\n淮阳区{}时，PM2.5浓度为{}μg/m3，在全市9个区县中排名第{}；PM10浓度为{}μg/m3，在全市9个区县中排名第{}。气象数据缺失。".format(
                                l[0][0:2], l[1], l[2], l[3], l[4]))
                    time.sleep(1)
                    fasong()
                    time.sleep(1)
                    CloseWindow(name)
                    print("已发送app")
                    time.sleep(2)
                    FindWindow(name)
                    table_opr.table_font("周口市九区县{}时PM2.5和PM10排名及各污染物详情表".format(l[0][0:2]))
                    table_opr.set_from_center()
                    excel_catch_screen()
                    print("已截图")
                    fasong()
                    time.sleep(1)
                    CloseWindow(name)
                    print("已发送截图")
                    time.sleep(2)
                    line_bar(l[0][0:2])
                    line_barpm10(l[0][0:2])
                    FindWindow(name)
                    for i in get_file("D:\Program Files\pycharm\机器人发送数据\image"):
                        paste_img("D:\Program Files\pycharm\机器人发送数据\image" + "\\" + i)
                        print(i)
                        fasong()
                    time.sleep(1)
                    print("柱状图已发送")
                    CloseWindow(name)
                    del_files("D:\Program Files\pycharm\机器人发送数据\image")
                    time.sleep(3550)
                else:
                    raise
            except:
                setText("数据异常！")
                time.sleep(1)
                fasong()
                CloseWindow(name)
                time.sleep(3570)
        except:
            time.sleep(60)