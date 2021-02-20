import openpyxl
import math
import time


name = input('请输入烟囱高度（m）：')
name1 = input('请输入烟囱直径（m）：')
name2 = input('请输入排放速率 (g/s)：')
name3 = input('请输入气体出口速度 (m/s)：')
name4 = input('请输入气体出口温度（℃）：')
name5 = input('请输入环境温度（℃）：')
name6 = input('请输入大气条件类别（1：非常不稳定；2：重度不稳定；3：轻微不稳定；4：中立；5：稳定；6：非常稳定）：')

# name = '30'
# name1 = '2'
# name2 = '5'
# name3 = '5'
# name4 = '200'
# name5 = '20'
# name6 = '5'

speed = [1,3,5,7,9,11,13,15,17,19]
dirc = [0,0.5,0.8,1.5,3,5,10,20,35,60,100]
sp = {'1':0.22,'2':0.16,'3':0.11,'4':0.08,'5':0.06,'6':0.04}
spp = {'1':0.2,'2':0.12,'3':0.08,'4':0.06,'5':0.03,'6':0.016}
sp1 = sp[name6]
spp1 = spp[name6]
# 定义温度

print('正在计算开尔文温度数据！')
time.sleep(1)
Kelvin = [int(name4)+273.15,int(name5)+273.15]
Kelvin.append(3.12*0.785*int(name3)*(int(name1)**2)*(Kelvin[0]-Kelvin[1])/(Kelvin[0]))
if Kelvin[2] > 55:
    Kelvin.append(34*(math.exp(0.4*math.log(Kelvin[2]))))
else:
    Kelvin.append(14*(math.exp(0.625*math.log(Kelvin[2]))))
Kelvin.append(1.6*math.exp(math.log(Kelvin[2])/3)*math.exp(math.log(3.5*Kelvin[3])*2/3))
if int(name6) < 6:
    Kelvin.append(9.806*0.02/Kelvin[1])
else:
    Kelvin.append(9.806 * 0.035 / Kelvin[1])
Lateral = []
Vertical = []
print('正在计算稳定性类别的横向和纵向离散系数！')
time.sleep(1)
for i in dirc:
    Lateral.append(sp1*1000*i*1/math.sqrt(1+0.1*i))

if 2>=int(name6)>=1:
    for j in dirc:
        Vertical.append(spp1*1000*j)
elif int(name6)==3:
    for j in dirc:
        Vertical.append(spp1*1000*j/(math.sqrt(1+0.2*j)))
elif int(name6)==4:
    for j in dirc:
        Vertical.append(spp1*1000*j/(math.sqrt(1+1.5*j)))
elif 6>=int(name6)>=5:
    for j in dirc:
        Vertical.append(spp1*1000*j/(1+0.3*j))

wb = openpyxl.load_workbook('data/demo.xlsx')
sheet = wb["Sheet1"]
sheet.cell(1, 1).value = 'speed'
sheet.cell(1, 2).value = '1'
sheet.cell(1, 3).value = '3'
sheet.cell(1, 4).value = '5'
sheet.cell(1, 5).value = '7'
sheet.cell(1, 6).value = '9'
sheet.cell(1, 7).value = '11'
sheet.cell(1, 8).value = '13'
sheet.cell(1, 9).value = '15'
sheet.cell(1, 10).value = '17'
sheet.cell(1, 11).value = '19'

sheet.cell(1, 3).value = '0'
sheet.cell(1, 4).value = '0.5'
sheet.cell(1, 5).value = '0.8'
sheet.cell(1, 6).value = '1.5'
sheet.cell(1, 7).value = '3'
sheet.cell(1, 8).value = '5'
sheet.cell(1, 9).value = '10'
sheet.cell(1, 10).value = '20'
sheet.cell(1, 11).value = '35'
sheet.cell(1, 12).value = '60'
sheet.cell(1, 13).value = '100'
# 写入Ht
sheet.cell(1, 2).value = 'Ht(m)'
Ht = []

for k in range(len(speed)):
    if int(name6) < 5:
        sl = int(name)+Kelvin[4]/speed[k]
    else:
        sl = int(name)+max(2.4*math.exp(math.log(Kelvin[2]/(Kelvin[5]*speed[k]))/3),5*math.exp(math.log(math.log(Kelvin[2]/(Kelvin[5]*speed[k]))/3)))
    sheet.cell(k+2, 2).value = sl
    sheet.cell(k + 2, 1).value = speed[k]
    Ht.append(sl)

print('正在计算烟雨中心线上距离原点的距离处污染物浓度！')
time.sleep(1)
# 写入落地浓度
for n in range(len(Lateral)):
    for m in range(len(speed)):
        if Lateral[n] < 0.01 or Vertical[n] < 0.01:
            sk = 0
        else:
            sk = 1000000*int(name2)/(2*math.pi*Lateral[n]*Vertical[n]*speed[m])*(math.exp(-0.5*((Ht[m]/Vertical[n])**2))*2)

        sheet.cell(m + 2, n+3).value = sk
wb.save('data/demo.xlsx')